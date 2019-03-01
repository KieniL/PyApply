from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os

#create new Excel Workboox
wbWrite=Workbook()

# set file path of output document
filepath = 'Output.xlsx'
#filepath = os.path.join('C:\\Users\\lkrei\\Desktop', filename)
wbWrite.save(filepath)


wsWrite1 = wbWrite.active
#set SheetTitle
wsWrite1.title = "Jobs"

#load Input Document
wb = load_workbook('Jobs.xlsx')
ws = wb['Tabelle1']

first_WritingColumn = 1
second_WritingColumn = 2
writeCounter = 1

first_column = ws['A']
second_column = ws['B']


options = Options()

#Add option that Browser is always maximized
options.add_argument("--start-maximized")

#Create webdriver on chromedriver from folder
browser = webdriver.Chrome('../../chromedriver.exe', options=options)

#load webpage
browser.get('https://www.karriere.at/jobs')


#Do for every Line in Input document
for x in range(len(first_column)):
    print("Jobs für "+ first_column[x].value + " in " + second_column[x].value)
    #find htmlElement for Job Description
    jobdesc = browser.find_element_by_name("keywords")
    #Clear Element if is already set (e.g. at Second Iteration)
    jobdesc.clear()
    #set Element Value from Input Document
    jobdesc.send_keys(first_column[x].value)
    #find htmlElement for Location
    location = browser.find_element_by_name("locations")
    #Clear Element if is already set (e.g. at Second Iteration)
    location.clear()
    #set Element Value from Input Document
    location.send_keys(second_column[x].value)

    #click search Button
    browser.find_element_by_class_name('m-jobsSearchform__submit').click()
    #wait for 3 seconds to page is loaded
    time.sleep(3)

    while True:
        #create array from htmlElements (get all elements with this className)
        jobs = browser.find_elements_by_class_name('m-jobItem__titleLink')
        #iterate over array
        for job in jobs:
            #write array into output
            wsWrite1.cell(row = writeCounter, column = first_WritingColumn).value = job.text
            wsWrite1.cell(row = writeCounter, column = second_WritingColumn).value = job.get_attribute("href")
            writeCounter += 1
        try:
            #click on next Button as long as possible
            browser.find_element_by_class_name("m-pagination__button--next").click()
            time.sleep(1)
        except Exception:
            print(first_column[x].value + " ist abgeschlossen")
            break

#save page after done all iterations
wbWrite.save(filepath)
#Finish Output
print("Gespeichert und erledigt. Datei liegt unter " + filepath)
